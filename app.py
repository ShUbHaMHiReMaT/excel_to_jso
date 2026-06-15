from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import json

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/convert", methods=["POST"])
def convert():

    if "file" not in request.files:
        return "No file uploaded"

    file = request.files["file"]

    if file.filename == "":
        return "No file selected"

    filename = secure_filename(file.filename)

    filepath = os.path.join(UPLOAD_FOLDER, filename)

    file.save(filepath)

    workbook = load_workbook(filepath)

    result = {}

    for sheet_name in workbook.sheetnames:

        sheet = workbook[sheet_name]

        rows = list(sheet.values)

        if not rows:
            continue

        headers = rows[0]

        data = []

        for row in rows[1:]:

            item = {}

            for header, value in zip(headers, row):
                item[str(header)] = value

            data.append(item)

        result[sheet_name] = data

    json_filename = filename.rsplit(".", 1)[0] + ".json"

    json_path = os.path.join(OUTPUT_FOLDER, json_filename)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=4, default=str)

    return render_template(
        "index.html",
        download_ready=True,
        download_file=json_filename
    )

@app.route("/download/<filename>")
def download(filename):
    return send_file(
        os.path.join(OUTPUT_FOLDER, filename),
        as_attachment=True
    )

if __name__ == "__main__":
    app.run()